import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, Listbox, MULTIPLE, END, Scrollbar, RIGHT, Y, LEFT, BOTH
import shutil
import os
import openpyxl
from openpyxl.styles import PatternFill

class CSVtoExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Analyzer")
        self.root.geometry("300x150")
        self.create_widgets()

    def create_widgets(self):
        self.load_csv_button = tk.Button(self.root, text="Load CSV Files", command=self.load_csv_files)
        self.load_csv_button.pack(pady=10)

        self.save_button = tk.Button(self.root, text="Save to Excel", command=self.save_to_excel)
        self.save_button.pack(pady=10)

        self.listbox_button = tk.Button(self.root, text="Select ComponentIDs", command=self.open_listbox)
        self.listbox_button.pack(pady=10)

    def load_csv_files(self):
        self.csv_files = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
        if self.csv_files:
            messagebox.showinfo("Loaded", f"Loaded {len(self.csv_files)} CSV files")
            self.file_path = self.csv_files[0]  # Store the first selected CSV file path

    def save_to_excel(self):
        if not hasattr(self, 'csv_files') or not self.csv_files:
            messagebox.showwarning("No Files", "No CSV files loaded")
            return
        
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        
        if file_path:
            combined_df = pd.DataFrame()
            for csv_file in self.csv_files:
                df = pd.read_csv(csv_file)
                file_name = os.path.splitext(os.path.basename(csv_file))[0]  # Extract file name without extension

                # Keep only the specified columns
                columns_to_keep = ['RESULT', 'ComponentID', 'VOLUME', 'HEIGHT', 'AREA', 'Panel']
                df = df[columns_to_keep]

                # Suffix the values in the "Panel" column with the file name
                if 'Panel' in df.columns:
                    df['Panel'] = df['Panel'].apply(lambda x: f"{x}_{file_name}" if pd.notnull(x) else x)

                # Remove the suffix from the "ComponentID" column
                if 'ComponentID' in df.columns:
                    df['ComponentID'] = df['ComponentID'].apply(lambda x: x.split('_')[0] if pd.notnull(x) else x)

                # Add the "Mils" column next to the "Height" column
                if 'HEIGHT' in df.columns:
                    df.insert(df.columns.get_loc('HEIGHT') + 1, 'Mils', df['HEIGHT'] / 25.4)

                combined_df = pd.concat([combined_df, df], ignore_index=True)

            self.combined_df = combined_df
            self.save_combined_df(file_path)

    def apply_conditional_formatting(self, writer, sheet_name):
        # Apply conditional formatting directly to the Excel file
        workbook = writer.book
        worksheet = workbook[sheet_name]
        for idx, row in self.combined_df.iterrows():
            result = row['RESULT']
            if result == 'GOOD':
                fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
            elif 'Warning' in result:
                fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light Blue
            elif 'Error' in result:
                fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
            else:
                fill = None  # No fill for other cases

            if fill:
                for cell in worksheet[f'A{idx + 2}:G{idx + 2}'][0]:  # Assuming columns A to G contain data
                    cell.fill = fill

    def apply_conditional_formatting_transposed(self, writer, sheet_name):
        # Apply conditional formatting to the transposed data
        workbook = writer.book
        worksheet = workbook[sheet_name]
        for idx, row in self.transposed_data_df.iterrows():
            component_id = row['ComponentID']
            for col in self.transposed_data_df.columns[1:]:
                cell_value = row[col]
                if pd.notnull(cell_value):
                    original_result = self.original_results[(self.original_results['ComponentID'] == component_id) & (self.original_results['Mils'] == cell_value)]['RESULT'].iloc[0]

                    if original_result == 'GOOD':
                        fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
                    elif 'Warning' in original_result:
                        fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light Blue
                    elif 'Error' in original_result:
                        fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
                    else:
                        fill = None  # No fill for other cases

                    if fill:
                        cell = worksheet.cell(row=idx + 2, column=self.transposed_data_df.columns.get_loc(col) + 1)
                        cell.fill = fill

    def save_combined_df(self, file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            self.combined_df.to_excel(writer, index=False, sheet_name="Combined Data")
            self.apply_conditional_formatting(writer, sheet_name="Combined Data")
            if hasattr(self, 'sampled_data_df'):
                self.sampled_data_df.to_excel(writer, index=False, sheet_name="Sampled Data")
            if hasattr(self, 'transposed_data_df'):
                self.transposed_data_df.to_excel(writer, index=False, sheet_name="Transposed Data")
                self.apply_conditional_formatting_transposed(writer, sheet_name="Transposed Data")
        messagebox.showinfo("Saved", f"Excel file saved to {file_path}")

    def open_listbox(self):
        if not hasattr(self, 'combined_df'):
            messagebox.showwarning("No Data", "No combined data available")
            return

        unique_component_ids = self.combined_df['ComponentID'].unique()

        self.listbox_window = tk.Toplevel(self.root)
        self.listbox_window.title("Select ComponentIDs")
        self.listbox_window.geometry("300x300")

        scrollbar = Scrollbar(self.listbox_window)
        scrollbar.pack(side=RIGHT, fill=Y)

        self.listbox = Listbox(self.listbox_window, selectmode=MULTIPLE, yscrollcommand=scrollbar.set)
        for item in unique_component_ids:
            self.listbox.insert(END, item)
        self.listbox.pack(side=LEFT, fill=BOTH, expand=True)
        
        scrollbar.config(command=self.listbox.yview)

        select_button = tk.Button(self.listbox_window, text="Select", command=self.get_selected_ids)
        select_button.pack(pady=10)

    def get_selected_ids(self):
        selected_indices = self.listbox.curselection()
        selected_ids = [self.listbox.get(i) for i in selected_indices]

        if len(selected_ids) > 100:  # Assuming you want a maximum of 100 ComponentIDs
            messagebox.showwarning("Invalid Selection", "Please select up to 100 ComponentIDs")
            return

        sampled_data = []
        self.original_results = pd.DataFrame()
        for comp_id in selected_ids:
            comp_data = self.combined_df[self.combined_df['ComponentID'] == comp_id].copy()
            if len(comp_data) > 20:
                comp_data = comp_data.sample(20)
            sampled_data.append(comp_data[['ComponentID', 'Mils']])
            self.original_results = pd.concat([self.original_results, comp_data[['ComponentID', 'Mils', 'RESULT']]], ignore_index=True)

        sampled_data_df = pd.concat(sampled_data, ignore_index=True)

        # Transpose the data
        transposed_data = {'ComponentID': selected_ids}
        max_samples = max(len(sampled_data_df[sampled_data_df['ComponentID'] == comp_id]) for comp_id in selected_ids)
        
        for i in range(max_samples):
            transposed_data[f'Sample_{i + 1}'] = [sampled_data_df[sampled_data_df['ComponentID'] == comp_id]['Mils'].iloc[i] if i < len(sampled_data_df[sampled_data_df['ComponentID'] == comp_id]) else None for comp_id in selected_ids]
        
        self.transposed_data_df = pd.DataFrame(transposed_data)
        
        self.listbox_window.destroy()
        messagebox.showinfo("Selected", f"Selected ComponentIDs: {', '.join(selected_ids)}")

        # Copy the target file to the working directory
        target_file_path = r"D:\NX_BACKWORK\Database_File\SMT_Data Analyzer\CPK-SPC-Xbar,R-chart.xlsx"
        working_directory = os.path.dirname(self.file_path)
        copied_file_path = os.path.join(working_directory, "CPK-SPC-Xbar,R-chart.xlsx")
        shutil.copy(target_file_path, copied_file_path)

if __name__ == "__main__":
    root = tk.Tk()
    app = CSVtoExcelApp(root)
    root.mainloop()



"""import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, Listbox, MULTIPLE, END, Scrollbar, RIGHT, Y, LEFT, BOTH
import shutil
import os

class CSVtoExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Analyzer")
        self.root.geometry("300x150")
        self.create_widgets()

    def create_widgets(self):
        self.load_csv_button = tk.Button(self.root, text="Load CSV Files", command=self.load_csv_files)
        self.load_csv_button.pack(pady=10)

        self.save_button = tk.Button(self.root, text="Save to Excel", command=self.save_to_excel)
        self.save_button.pack(pady=10)

        self.listbox_button = tk.Button(self.root, text="Select ComponentIDs", command=self.open_listbox)
        self.listbox_button.pack(pady=10)

    def load_csv_files(self):
        self.csv_files = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
        if self.csv_files:
            messagebox.showinfo("Loaded", f"Loaded {len(self.csv_files)} CSV files")
            # Store the file path
            self.file_path = self.csv_files[0]

    def save_to_excel(self):
        if not hasattr(self, 'csv_files') or not self.csv_files:
            messagebox.showwarning("No Files", "No CSV files loaded")
            return
        
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        
        if file_path:
            combined_df = pd.DataFrame()
            for csv_file in self.csv_files:
                df = pd.read_csv(csv_file)
                file_name = csv_file.split('/')[-1].split('.')[0]

                # Keep only the specified columns
                columns_to_keep = ['RESULT', 'ComponentID', 'VOLUME', 'HEIGHT', 'AREA', 'Panel']
                df = df[columns_to_keep]

                # Suffix the values in the "Panel" column with the file name
                if 'Panel' in df.columns:
                    df['Panel'] = df['Panel'].apply(lambda x: f"{x}_{file_name}" if pd.notnull(x) else x)

                # Remove the suffix from the "ComponentID" column
                if 'ComponentID' in df.columns:
                    df['ComponentID'] = df['ComponentID'].apply(lambda x: x.split('_')[0] if pd.notnull(x) else x)

                # Add the "Mils" column next to the "Height" column
                if 'HEIGHT' in df.columns:
                    df.insert(df.columns.get_loc('HEIGHT') + 1, 'Mils', df['HEIGHT'] / 25.4)

                combined_df = pd.concat([combined_df, df], ignore_index=True)

            self.combined_df = combined_df
            self.save_combined_df(file_path)

    def save_combined_df(self, file_path):
        with pd.ExcelWriter(file_path) as writer:
            self.combined_df.to_excel(writer, index=False, sheet_name="Combined Data")
            if hasattr(self, 'sampled_data_df'):
                self.sampled_data_df.to_excel(writer, index=False, sheet_name="Sampled Data")
            if hasattr(self, 'transposed_data_df'):
                self.transposed_data_df.to_excel(writer, index=False, sheet_name="Transposed Data")
        messagebox.showinfo("Saved", f"Excel file saved to {file_path}")

    def open_listbox(self):
        if not hasattr(self, 'combined_df'):
            messagebox.showwarning("No Data", "No combined data available")
            return

        unique_component_ids = self.combined_df['ComponentID'].unique()

        self.listbox_window = tk.Toplevel(self.root)
        self.listbox_window.title("Select ComponentIDs")
        self.listbox_window.geometry("300x300")

        scrollbar = Scrollbar(self.listbox_window)
        scrollbar.pack(side=RIGHT, fill=Y)

        self.listbox = Listbox(self.listbox_window, selectmode=MULTIPLE, yscrollcommand=scrollbar.set)
        for item in unique_component_ids:
            self.listbox.insert(END, item)
        self.listbox.pack(side=LEFT, fill=BOTH, expand=True)
        
        scrollbar.config(command=self.listbox.yview)

        select_button = tk.Button(self.listbox_window, text="Select", command=self.get_selected_ids)
        select_button.pack(pady=10)

    def get_selected_ids(self):
        selected_indices = self.listbox.curselection()
        selected_ids = [self.listbox.get(i) for i in selected_indices]

        if len(selected_ids) >100: #!= 5:
            messagebox.showwarning("Invalid Selection", "Please select exactly 5 ComponentIDs")
            return

        sampled_data = []
        for comp_id in selected_ids:
            comp_data = self.combined_df[self.combined_df['ComponentID'] == comp_id].copy()
            if len(comp_data) > 20:
                comp_data = comp_data.sample(20)
            sampled_data.append(comp_data[['ComponentID', 'Mils']])
        
        sampled_data_df = pd.concat(sampled_data, ignore_index=True)

        # Transpose the data
        transposed_data = {'ComponentID': selected_ids}
        max_samples = max(len(sampled_data_df[sampled_data_df['ComponentID'] == comp_id]) for comp_id in selected_ids)
        
        for i in range(max_samples):
            transposed_data[f'Sample_{i + 1}'] = [sampled_data_df[sampled_data_df['ComponentID'] == comp_id]['Mils'].iloc[i] if i < len(sampled_data_df[sampled_data_df['ComponentID'] == comp_id]) else None for comp_id in selected_ids]
        
        self.transposed_data_df = pd.DataFrame(transposed_data)
        
        self.listbox_window.destroy()
        messagebox.showinfo("Selected", f"Selected ComponentIDs: {', '.join(selected_ids)}")

        # Copy the target file to the working directory
        target_file_path = r"D:\NX_BACKWORK\Database_File\SMT_Data Analyzer\CPK-SPC-Xbar,R-chart.xlsx"
        working_directory = os.path.dirname(self.file_path)
        copied_file_path = os.path.join(working_directory, "CPK-SPC-Xbar,R-chart.xlsx")
        shutil.copy(target_file_path, copied_file_path)

if __name__ == "__main__":
    root = tk.Tk()
    app = CSVtoExcelApp(root)
    root.mainloop()"""