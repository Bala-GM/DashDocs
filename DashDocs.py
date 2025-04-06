import os
import time
import tabula
import shutil
import subprocess
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import simpledialog, messagebox
from tkinter import filedialog, messagebox, Listbox, MULTIPLE, END, Scrollbar, RIGHT, Y, LEFT, BOTH
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill


#=====================================================================================================================

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("DashDocs")
        self.root.geometry("300x200")

        self.program_var = tk.StringVar(value="Program1")

        # Radio buttons to choose the program
        tk.Label(root, text="Choose a Program: PY V-2.0.0 APR|06|04|2025").pack(pady=5)
        tk.Radiobutton(root, text="Program 1-AOI", variable=self.program_var, value="Program1").pack(anchor=tk.W)
        tk.Radiobutton(root, text="Program 2-SPI", variable=self.program_var, value="Program2").pack(anchor=tk.W)
        tk.Radiobutton(root, text="Program 3-RO-single", variable=self.program_var, value="Program3").pack(anchor=tk.W)
        tk.Radiobutton(root, text="Program 4-RO-Multi", variable=self.program_var, value="Program4").pack(anchor=tk.W)
        tk.Label(root, text="MIT License\n\nCopyright (C) <2025>  <BALA GANESH>").pack(pady=5)

        # Button to run the selected program
        self.run_button = tk.Button(root, text="Run Program", command=self.run_selected_program)
        self.run_button.pack(pady=20)

    def run_selected_program(self):
        if self.program_var.get() == "Program1":
            self.run_program1()
        elif self.program_var.get() == "Program2":
            self.run_program2()
        elif self.program_var.get() == "Program3":
            self.run_program3()
        elif self.program_var.get() == "Program4":
            self.run_program4()

    def run_program1(self):
        # Code for Program 1
        self.new_window = tk.Toplevel(self.root)
        self.app1 = ExcelProcessorAppProgram1(self.new_window)

    def run_program2(self):
        # Code for Program 2
        self.new_window = tk.Toplevel(self.root)
        self.app2 = CSVtoExcelApp(self.new_window)
    
    def run_program3(self):
        # Code for Program 3
        self.new_window = tk.Toplevel(self.root)
        self.app3 = PDFtoExcelApp1(self.new_window)
    
    def run_program4(self):
        # Code for Program 4
        self.new_window = tk.Toplevel(self.root)
        self.app4 = PDFtoExcelApp2(self.new_window)

#===============================================================================================================

class ExcelProcessorAppProgram1:
    def __init__(self, root):
        self.root = root
        self.root.title("AOI Kohyoung") #Excel Processor
        self.root.geometry("250x100")

        self.upload_button = tk.Button(root, text="Upload Excel File", command=self.upload_file)
        self.upload_button.pack(pady=10)

        self.process_button = tk.Button(root, text="Process and Save Excel", command=self.process_and_save)
        self.process_button.pack(pady=10)

    def upload_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.file_path = file_path
        messagebox.showinfo("File Upload", f"Uploaded file: {self.file_path}")

    def process_and_save(self):
        if not self.file_path:
            messagebox.showwarning("No File", "No file has been uploaded.")
            return

        # Load the Excel file
        wb = load_workbook(self.file_path)

        # Rename the first sheet to FRM2
        first_sheet_name = wb.sheetnames[0]
        wb[first_sheet_name].title = "FRM2"

        # Create a new sheet FRM1 and copy the first two rows from FRM2
        frm2_sheet = wb["FRM2"]
        if "FRM1" in wb.sheetnames:
            frm1_sheet = wb["FRM1"]
        else:
            frm1_sheet = wb.create_sheet("FRM1")
        for row in frm2_sheet.iter_rows(max_row=2):
            frm1_sheet.append([cell.value for cell in row])

        # Delete the first three rows from FRM2
        frm2_sheet.delete_rows(1, 3)

        # Check and delete the columns Assign Level and Inspection Name from FRM2
        header_columns = ['Assign Level', 'Inspection Name']
        for col in header_columns:
            for row in frm2_sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    if cell.value == col:
                        frm2_sheet.delete_cols(cell.column)

        # Combine data from FRM1 and FRM2 into a new sheet FRM3 with a one-row gap
        new_combined_sheet = wb.create_sheet("FRM3")

        # Copy data from FRM1 to FRM3
        for row in frm1_sheet.iter_rows(values_only=True):
            new_combined_sheet.append(row)

        # Add an empty row for the gap
        new_combined_sheet.append([None] * frm1_sheet.max_column)

        # Copy data from FRM2 to FRM3 after the gap
        for row in frm2_sheet.iter_rows(values_only=True):
            new_combined_sheet.append(row)

        # Delete sheets FRM1 and FRM2
        wb.remove(frm1_sheet)
        wb.remove(frm2_sheet)

        # Save the modified Excel file temporarily
        temp_output_file = self.file_path.replace(".xlsx", "_temp.xlsx")
        wb.save(temp_output_file)

        # Copy the target file to the working directory
        target_file_path = r"D:\NX_BACKWORK\Database_File\SMT_Data Analyzer\2321 AOI COVERAGE REPORT REV E.xlsx"
        working_directory = os.path.dirname(self.file_path)
        copied_file_path = os.path.join(working_directory, "2321 AOI COVERAGE REPORT REV E.xlsx")
        shutil.copy(target_file_path, copied_file_path)

        # Open the copied target file and paste the data from FRM3 into the Kohyoung sheet
        target_wb = load_workbook(copied_file_path)
        kohyoung_sheet = target_wb["Kohyoung"]
        frm3_sheet = wb["FRM3"]

        # Paste data starting from A2 in Kohyoung sheet
        start_row = 2
        for row in frm3_sheet.iter_rows(values_only=True):
            for col_idx, cell in enumerate(row, start=1):
                kohyoung_sheet.cell(row=start_row, column=col_idx, value=cell)
            start_row += 1

        # Hide blank rows in the Kohyoung sheet
        for row in kohyoung_sheet.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                kohyoung_sheet.row_dimensions[row[0].row].hidden = True

        # Save the final output file
        final_output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not final_output_file:
            return

        # Save the target workbook
        target_wb.save(final_output_file)

        # Remove the temporary file
        os.remove(temp_output_file)

        messagebox.showinfo("Data Saved", f"Data has been processed and saved to {final_output_file}")

        # Attempt to delete the file with a delay and retries
        for _ in range(3):  # Try a few times
            try:
                os.remove(copied_file_path)
                break  # Exit the loop if deletion is successful
            except PermissionError:
                # If PermissionError occurs, wait for a short time before retrying
                time.sleep(0.5)  # Wait for 0.5 seconds before retrying
        else:
            # If deletion fails after several attempts, display a message to the user
            messagebox.showwarning("Deletion Error", "Failed to delete the file. Please close any applications using the file.")

#=====================================================================================================================

class CSVtoExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("SPI Data Analyzer")
        self.root.geometry("300x150")
        self.create_widgets()

    def create_widgets(self):
        self.load_csv_button = tk.Button(self.root, text="Load CSV Files", command=self.load_csv_files)
        self.load_csv_button.pack(pady=10)

        self.save_button = tk.Button(self.root, text="Save to Excel", command=self.save_to_excel)
        self.save_button.pack(pady=10)

        self.listbox_button = tk.Button(self.root, text="Select ComponentIDs", command=self.open_listbox)
        self.listbox_button.pack(pady=10)

    def load_csv_files(self):
        self.csv_files = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
        if self.csv_files:
            messagebox.showinfo("Loaded", f"Loaded {len(self.csv_files)} CSV files")
            self.file_path = self.csv_files[0]  # Store the first selected CSV file path

    def save_to_excel(self):
        if not hasattr(self, 'csv_files') or not self.csv_files:
            messagebox.showwarning("No Files", "No CSV files loaded")
            return
        
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        
        if file_path:
            combined_df = pd.DataFrame()
            for csv_file in self.csv_files:
                df = pd.read_csv(csv_file)
                file_name = os.path.splitext(os.path.basename(csv_file))[0]  # Extract file name without extension

                # Keep only the specified columns
                columns_to_keep = ['RESULT', 'ComponentID', 'VOLUME', 'HEIGHT', 'AREA', 'Panel']
                df = df[columns_to_keep]

                # Suffix the values in the "Panel" column with the file name
                if 'Panel' in df.columns:
                    df['Panel'] = df['Panel'].apply(lambda x: f"{x}_{file_name}" if pd.notnull(x) else x)

                # Remove the suffix from the "ComponentID" column
                if 'ComponentID' in df.columns:
                    df['ComponentID'] = df['ComponentID'].apply(lambda x: x.split('_')[0] if pd.notnull(x) else x)

                # Add the "Mils" column next to the "Height" column
                if 'HEIGHT' in df.columns:
                    df.insert(df.columns.get_loc('HEIGHT') + 1, 'Mils', df['HEIGHT'] / 25.4)

                combined_df = pd.concat([combined_df, df], ignore_index=True)

            self.combined_df = combined_df
            self.save_combined_df(file_path)

    def apply_conditional_formatting(self, writer, sheet_name):
        # Apply conditional formatting directly to the Excel file
        workbook = writer.book
        worksheet = workbook[sheet_name]
        for idx, row in self.combined_df.iterrows():
            result = row['RESULT']
            if result == 'GOOD':
                fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
            elif 'Warning' in result:
                fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light Blue
            elif 'Error' in result:
                fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
            else:
                fill = None  # No fill for other cases

            if fill:
                for cell in worksheet[f'A{idx + 2}:G{idx + 2}'][0]:  # Assuming columns A to G contain data
                    cell.fill = fill

    def apply_conditional_formatting_transposed(self, writer, sheet_name):
        # Apply conditional formatting to the transposed data
        workbook = writer.book
        worksheet = workbook[sheet_name]
        for idx, row in self.transposed_data_df.iterrows():
            component_id = row['ComponentID']
            for col in self.transposed_data_df.columns[1:]:
                cell_value = row[col]
                if pd.notnull(cell_value):
                    original_result = self.original_results[(self.original_results['ComponentID'] == component_id) & (self.original_results['Mils'] == cell_value)]['RESULT'].iloc[0]

                    if original_result == 'GOOD':
                        fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
                    elif 'Warning' in original_result:
                        fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light Blue
                    elif 'Error' in original_result:
                        fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
                    else:
                        fill = None  # No fill for other cases

                    if fill:
                        cell = worksheet.cell(row=idx + 2, column=self.transposed_data_df.columns.get_loc(col) + 1)
                        cell.fill = fill

    def save_combined_df(self, file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            self.combined_df.to_excel(writer, index=False, sheet_name="Combined Data")
            self.apply_conditional_formatting(writer, sheet_name="Combined Data")
            if hasattr(self, 'sampled_data_df'):
                self.sampled_data_df.to_excel(writer, index=False, sheet_name="Sampled Data")
            if hasattr(self, 'transposed_data_df'):
                self.transposed_data_df.to_excel(writer, index=False, sheet_name="Transposed Data")
                self.apply_conditional_formatting_transposed(writer, sheet_name="Transposed Data")
        messagebox.showinfo("Saved", f"Excel file saved to {file_path}")

    def open_listbox(self):
        if not hasattr(self, 'combined_df'):
            messagebox.showwarning("No Data", "No combined data available")
            return

        unique_component_ids = self.combined_df['ComponentID'].unique()

        self.listbox_window = tk.Toplevel(self.root)
        self.listbox_window.title("Select ComponentIDs")
        self.listbox_window.geometry("300x300")

        scrollbar = Scrollbar(self.listbox_window)
        scrollbar.pack(side=RIGHT, fill=Y)

        self.listbox = Listbox(self.listbox_window, selectmode=MULTIPLE, yscrollcommand=scrollbar.set)
        for item in unique_component_ids:
            self.listbox.insert(END, item)
        self.listbox.pack(side=LEFT, fill=BOTH, expand=True)
        
        scrollbar.config(command=self.listbox.yview)

        select_button = tk.Button(self.listbox_window, text="Select", command=self.get_selected_ids)
        select_button.pack(pady=10)

    def get_selected_ids(self):
        selected_indices = self.listbox.curselection()
        selected_ids = [self.listbox.get(i) for i in selected_indices]

        if len(selected_ids) > 100:  # Assuming you want a maximum of 100 ComponentIDs
            messagebox.showwarning("Invalid Selection", "Please select up to 100 ComponentIDs")
            return
        
        # Step 1: Ask for Stencil Mil
        root = tk.Tk()
        root.withdraw()

        stencil_mil = simpledialog.askfloat("Stencil Mil Input", "Enter Stencil Mil value:")
        if stencil_mil is None:
            messagebox.showinfo("Process Cancelled", "No input given. Operation aborted.")
            return  # or exit from your function

        # Step 2: Default Spec Limits
        upper_limit = stencil_mil + 2
        lower_limit = stencil_mil - 1

        # Step 3: Ask user if they want to tighten the spec
        tighten = messagebox.askyesno("Tighten Spec?", 
            f"Default Spec Limits:\nUpper: {upper_limit}\nLower: {lower_limit}\n\nWould you like to adjust them?")

        # Step 4: If yes, allow user to edit them
        if tighten:
            user_upper = simpledialog.askfloat("Edit Upper Limit", f"Enter custom Upper Limit (Default: {upper_limit}):")
            user_lower = simpledialog.askfloat("Edit Lower Limit", f"Enter custom Lower Limit (Default: {lower_limit}):")

            if user_upper is not None:
                upper_limit = user_upper
            if user_lower is not None:
                lower_limit = user_lower

        # Step 5: Final confirmation
        messagebox.showinfo("Final Spec Limits", 
            f"Stencil Mil: {stencil_mil}\nUpper Limit: {upper_limit}\nLower Limit: {lower_limit}")

        # --- Continue with your sampling code ---

        sampled_data = []
        self.original_results = pd.DataFrame()

        for comp_id in selected_ids:
            comp_data = self.combined_df[self.combined_df['ComponentID'] == comp_id].copy()
            if len(comp_data) > 100:
                comp_data = comp_data.sample(100)

            # Filter within spec
            comp_data = comp_data[(comp_data['Mils'] >= lower_limit) & (comp_data['Mils'] <= upper_limit)]

            sampled_data.append(comp_data[['ComponentID', 'Mils']])
            self.original_results = pd.concat([self.original_results, comp_data[['ComponentID', 'Mils', 'RESULT', 'VOLUME', 'HEIGHT', 'AREA', 'Panel']]], ignore_index=True)

        sampled_data_df = pd.concat(sampled_data, ignore_index=True)

        # Save out-of-spec data to another sheet
        out_of_spec_df = self.combined_df[(self.combined_df['ComponentID'].isin(selected_ids)) &
                                        ((self.combined_df['Mils'] < lower_limit) | (self.combined_df['Mils'] > upper_limit))]

        if not out_of_spec_df.empty:
            out_of_spec_path = os.path.join(os.path.dirname(self.file_path), "Out_of_Spec_Data.xlsx")
            out_of_spec_df.to_excel(out_of_spec_path, index=False)

        # Transpose the data
        transposed_data = {'ComponentID': selected_ids}
        max_samples = max(len(sampled_data_df[sampled_data_df['ComponentID'] == comp_id]) for comp_id in selected_ids)

        for i in range(max_samples):
            transposed_data[f'Sample_{i + 1}'] = [sampled_data_df[sampled_data_df['ComponentID'] == comp_id]['Mils'].iloc[i] if i < len(sampled_data_df[sampled_data_df['ComponentID'] == comp_id]) else None for comp_id in selected_ids]

        self.transposed_data_df = pd.DataFrame(transposed_data)

        self.listbox_window.destroy()
        messagebox.showinfo("Selected", f"Selected ComponentIDs: {', '.join(selected_ids)}")

        # Copy the target file to the working directory
        target_file_path = r"D:\\NX_BACKWORK\\Database_File\\SMT_Data Analyzer\\CPK-SPC-Xbar,R-chart.xlsx"
        working_directory = os.path.dirname(self.file_path)
        copied_file_path = os.path.join(working_directory, "CPK-SPC-Xbar,R-chart.xlsx")
        shutil.copy(target_file_path, copied_file_path)


#====================================================================================================================

print(os.environ.get('JAVA_HOME'))

try:
    result = subprocess.run(['java', '-version'], capture_output=True, text=True)
    print(result.stdout)
except FileNotFoundError:
    print("Java command not found. Ensure JAVA_HOME is set correctly.")


class PDFtoExcelApp1:
    def __init__(self, root):
        self.root = root
        self.root.title("Reflow Data #single") #PDF to Excel Converter
        self.root.geometry("275x100")

        self.upload_button = tk.Button(root, text="Upload PDF File", command=self.upload_file)
        self.upload_button.pack(pady=10)
        
        self.process_button = tk.Button(root, text="Convert to Excel", command=self.process_file)
        self.process_button.pack(pady=10)
        
        self.file_path = None
    
    def upload_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if file_path:
            self.file_path = file_path
            messagebox.showinfo("File Upload", f"Uploaded file: {self.file_path}")
    
    def process_file(self):
        if not self.file_path:
            messagebox.showwarning("No File", "No file has been uploaded.")
            return
        
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if output_path:
            self.convert_pdf_to_excel(self.file_path, output_path)
            messagebox.showinfo("File Saved", f"Excel file saved to: {output_path}")

    def convert_pdf_to_excel(self, pdf_path, excel_path):
        # Extract tables using Tabula-py
        tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
        
        # Write each table to a separate sheet in the Excel file
        with pd.ExcelWriter(excel_path) as writer:
            for i, table in enumerate(tables):
                df = pd.DataFrame(table)
                df.to_excel(writer, index=False, sheet_name=f"Page_{i+1}_Table_{i+1}")

        # Copy the target file to the working directory
        target_file_path = r"D:\NX_BACKWORK\Database_File\SMT_Data Analyzer\Reflow Process Results - Cpk Calculator - Reflow Parameters - HrP2.xlsx"
        working_directory = os.path.dirname(self.file_path)
        copied_file_path = os.path.join(working_directory, "Reflow Process Results - Cpk Calculator - Reflow Parameters - HrP2.xlsx")
        shutil.copy(target_file_path, copied_file_path)

#=================================================================================================================

class PDFtoExcelApp2:
    def __init__(self, root):
        self.root = root
        self.root.title("Reflow Data #multi") #PDF to Excel Converter
        self.root.geometry("275x100")
        
        self.upload_button = tk.Button(root, text="Upload PDF Files", command=self.upload_files)
        self.upload_button.pack(pady=10)
        
        self.process_button = tk.Button(root, text="Convert to Excel", command=self.process_files)
        self.process_button.pack(pady=10)
        
        self.file_paths = []  # List to store selected file paths
    
    def upload_files(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        if file_paths:
            self.file_paths = file_paths
            messagebox.showinfo("Files Upload", f"Uploaded {len(self.file_paths)} files")
    
    def process_files(self):
        if not self.file_paths:
            messagebox.showwarning("No Files", "No files have been uploaded.")
            return
        
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if output_path:
            self.copy_target_excel(output_path)
            self.convert_pdfs_to_excel(self.file_paths, output_path)
            messagebox.showinfo("File Saved", f"Excel file saved to: {output_path}")

    def copy_target_excel(self, output_path):
        target_file_path = r"D:\NX_BACKWORK\Database_File\SMT_Data Analyzer\Reflow Process Results - Cpk Calculator - Reflow Parameters - HrP2.xlsx"
        working_directory = os.path.dirname(output_path)
        copied_file_path = os.path.join(working_directory, "Reflow Process Results - Cpk Calculator - Reflow Parameters - HrP2.xlsx")
        shutil.copy(target_file_path, copied_file_path)

    def convert_pdfs_to_excel(self, pdf_paths, excel_path):
        # Initialize an empty workbook
        wb = Workbook()
        
        # Create a sheet named "Combined"
        ws = wb.active
        ws.title = "Combined"
        
        # Write each table from PDFs to the "Combined" sheet
        for pdf_path in pdf_paths:
            tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
            if len(tables) >= 3:
                df = pd.DataFrame(tables[2])  # Assuming third table is at index 2
                
                # Add filename as a column
                df['Filename'] = pdf_path
                
                # Convert DataFrame rows to tuples and append to worksheet
                for row in dataframe_to_rows(df, index=False, header=True):
                    ws.append(row)
        
        # Save the workbook
        wb.save(excel_path)

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()

#pyinstaller -F -i "DD.ico" --onefile --noconsole DashDocs.py