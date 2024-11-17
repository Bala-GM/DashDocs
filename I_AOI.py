import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import shutil
import os

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Processor")
        self.root.geometry("200x150")

        self.upload_button = tk.Button(root, text="Upload Excel File", command=self.upload_file)
        self.upload_button.pack(pady=10)

        self.process_button = tk.Button(root, text="Process and Save Excel", command=self.process_and_save)
        self.process_button.pack(pady=10)

    def upload_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.file_path = file_path
        messagebox.showinfo("File Upload", f"Uploaded file: {self.file_path}")

    def process_and_save(self):
        if not self.file_path:
            messagebox.showwarning("No File", "No file has been uploaded.")
            return

        # Load the Excel file
        wb = load_workbook(self.file_path)

        # Rename the first sheet to FRM2
        first_sheet_name = wb.sheetnames[0]
        wb[first_sheet_name].title = "FRM2"

        # Create a new sheet FRM1 and copy the first two rows from FRM2
        frm2_sheet = wb["FRM2"]
        if "FRM1" in wb.sheetnames:
            frm1_sheet = wb["FRM1"]
        else:
            frm1_sheet = wb.create_sheet("FRM1")
        for row in frm2_sheet.iter_rows(max_row=2):
            frm1_sheet.append([cell.value for cell in row])

        # Delete the first three rows from FRM2
        frm2_sheet.delete_rows(1, 3)

        # Check and delete the columns Assign Level and Inspection Name from FRM2
        header_columns = ['Assign Level', 'Inspection Name']
        for col in header_columns:
            for row in frm2_sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    if cell.value == col:
                        frm2_sheet.delete_cols(cell.column)

        # Combine data from FRM1 and FRM2 into a new sheet FRM3 with a one-row gap
        new_combined_sheet = wb.create_sheet("FRM3")

        # Copy data from FRM1 to FRM3
        for row in frm1_sheet.iter_rows(values_only=True):
            new_combined_sheet.append(row)

        # Add an empty row for the gap
        new_combined_sheet.append([None] * frm1_sheet.max_column)

        # Copy data from FRM2 to FRM3 after the gap
        for row in frm2_sheet.iter_rows(values_only=True):
            new_combined_sheet.append(row)

        # Delete sheets FRM1 and FRM2
        wb.remove(frm1_sheet)
        wb.remove(frm2_sheet)

        # Save the modified Excel file temporarily
        temp_output_file = self.file_path.replace(".xlsx", "_temp.xlsx")
        wb.save(temp_output_file)

        # Copy the target file to the working directory
        target_file_path = r"D:\NX_BACKWORK\Database_File\SMT_Data Analyzer\2321 AOI COVERAGE REPORT REV E.xlsx"
        working_directory = os.path.dirname(self.file_path)
        copied_file_path = os.path.join(working_directory, "2321 AOI COVERAGE REPORT REV E.xlsx")
        shutil.copy(target_file_path, copied_file_path)

        # Open the copied target file and paste the data from FRM3 into the Kohyoung sheet
        target_wb = load_workbook(copied_file_path)
        kohyoung_sheet = target_wb["Kohyoung"]
        frm3_sheet = wb["FRM3"]

        # Paste data starting from A2 in Kohyoung sheet
        start_row = 2
        for row in frm3_sheet.iter_rows(values_only=True):
            for col_idx, cell in enumerate(row, start=1):
                kohyoung_sheet.cell(row=start_row, column=col_idx, value=cell)
            start_row += 1

        # Hide blank rows in the Kohyoung sheet
        for row in kohyoung_sheet.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                kohyoung_sheet.row_dimensions[row[0].row].hidden = True

        # Save the final output file
        final_output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not final_output_file:
            return

        # Save the target workbook
        target_wb.save(final_output_file)

        # Remove the temporary file
        os.remove(temp_output_file)

        messagebox.showinfo("Data Saved", f"Data has been processed and saved to {final_output_file}")

        import time

        # Attempt to delete the file with a delay and retries
        for _ in range(3):  # Try a few times
            try:
                os.remove(copied_file_path)
                break  # Exit the loop if deletion is successful
            except PermissionError:
                # If PermissionError occurs, wait for a short time before retrying
                time.sleep(0.5)  # Wait for 0.5 seconds before retrying
        else:
            # If deletion fails after several attempts, display a message to the user
            messagebox.showwarning("Deletion Error", "Failed to delete the file. Please close any applications using the file.")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()
